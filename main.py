import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook

# Function to search for the data
def search_data():
    id_search = entry_id.get()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Convert both to string for an exact match
        if str(row[0]).zfill(3) == id_search.zfill(3):
            entry_name.delete(0, tk.END)
            entry_name.insert(0, row[1])
            entry_usd.delete(0, tk.END)
            entry_usd.insert(0, str(row[2]) if row[2] is not None else "") #Update to str when search   
            entry_riel.delete(0, tk.END)
            entry_riel.insert(0, str(row[3]) if row[3] is not None else "") #Update to str whem search
            break


# Function to clear all the entry widgets
def clear_entries():
    entry_id.delete(0, tk.END)
    entry_name.delete(0, tk.END)
    entry_usd.delete(0, tk.END)
    entry_riel.delete(0, tk.END)

# Function to update the data
def update_data():
    try:
        id_search = entry_id.get()
        found = False
        for row in sheet.iter_rows(min_row=2):
            if str(row[0].value).zfill(3) == id_search.zfill(3):
                row[2].value = int(entry_usd.get())  # Assuming USD is a number
                row[3].value = int(entry_riel.get())  # Assuming Riel is text
                found = True
                break
        if found:
            wb.save('D:\Wed\data.xlsx')
            print("Data updated successfully.")
            refresh_data() # Refresh the Treeview with the updated data
        else:
            print("ID not found. Please check the ID and try again.")
    except ValueError:
        print("Please enter a valid number for USD.")
    except Exception as e:
        print(f"An error occurred: {e}")
    


# Function to load data from Excel into the Treeview
def load_data():
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tree.insert("", tk.END, values=row)


# Function to refresh the Treeview with updated data
def refresh_data():
    for item in tree.get_children():
        tree.delete(item)
    load_data()
    
 # Load the workbook and select the active sheet
wb = load_workbook('D:\\Wed\\data.xlsx')
sheet = wb.active   

# Load the workbook and select the active sheet
wb = load_workbook('D:\Wed\data.xlsx')
sheet = wb.active

# Create the main window
root = tk.Tk()
root.title("Wedding Guest Book")



# Create the entry widgets
entry_id = tk.Entry(root, width=40)
entry_id.grid(row=0, column=1)
entry_name = tk.Entry(root, width=40)
entry_name.grid(row=1, column=1)
entry_usd = tk.Entry(root, width=40)
entry_usd.grid(row=2, column=1)
entry_riel = tk.Entry(root, width=40)
entry_riel.grid(row=3, column=1)

# Create the label widgets
label_id = tk.Label(root, text="ID")
label_id.grid(row=0, column=0)
label_name = tk.Label(root, text="Name")
label_name.grid(row=1, column=0)
label_usd = tk.Label(root, text="USD")
label_usd.grid(row=2, column=0)
label_riel = tk.Label(root, text="Riel")
label_riel.grid(row=3, column=0)

# Create the button widgets
button_search = tk.Button(root, text="Search", command=search_data)
button_search.grid(row=4, column=0)
# Create the 'Clear' button and place it on the grid
button_clear = tk.Button(root, text="Clear", command=clear_entries)
button_clear.grid(row=5, column=1)
button_update = tk.Button(root, text="Update", command=update_data)
button_update.grid(row=4, column=1)

# Create the Treeview widget
tree = ttk.Treeview(root, columns=("ID", "Name", "USD", "Riel"), show='headings')
tree.heading("ID", text="ID")
tree.heading("Name", text="Name")
tree.heading("USD", text="USD")
tree.heading("Riel", text="Riel")
tree.grid(row=6, columnspan=4, sticky='nsew')

# Load data into the Treeview
load_data()

# Run the main loop
root.mainloop()
